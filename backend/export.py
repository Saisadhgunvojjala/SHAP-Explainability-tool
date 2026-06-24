"""
backend/export.py — Batch Excel export of SHAP explanations.

Generates an Excel workbook with three sheets:
  1. Predictions     — original data + prediction column
  2. SHAP Values     — raw SHAP value per feature per row
  3. Top Explanations — plain English top-3 reasons per row (non-technical)

Why three sheets?
  Technical users want raw SHAP values (sheet 2).
  Non-technical clients want plain English (sheet 3).
  Everyone needs the predictions (sheet 1).
"""

import io
import numpy as np
import pandas as pd


def _direction(val: float) -> str:
    return "increases prediction" if val >= 0 else "decreases prediction"


def _plain_english_row(row_idx: int, local_shap: np.ndarray,
                        feature_names: list, raw_row: dict,
                        prediction, top_n: int = 3) -> dict:
    """
    Generate plain-English top-N explanation for one row.
    Returns a flat dict suitable for a DataFrame row.
    """
    arr      = np.array(local_shap).flatten()
    abs_arr  = np.abs(arr)
    total    = abs_arr.sum()
    top_idx  = np.argsort(abs_arr)[::-1][:top_n]

    result = {"Row": row_idx + 1, "Prediction": prediction}

    for rank, idx in enumerate(top_idx):
        idx   = int(idx)
        feat  = feature_names[idx]
        sv    = arr[idx]
        pct   = (abs(sv) / total * 100) if total > 0 else 0
        val   = raw_row.get(feat, "N/A")
        result[f"Reason_{rank+1}"] = (
            f"{feat} = {val} "
            f"({_direction(sv)}, {pct:.1f}% impact, SHAP={sv:+.4f})"
        )

    return result


def generate_excel_export(
    raw_df:        pd.DataFrame,
    predictions:   np.ndarray,
    shap_values:   np.ndarray,
    feature_names: list,
    target_name:   str,
    shap_row_count: int,
) -> bytes:
    """
    Build and return Excel workbook as bytes.

    Args:
        raw_df:         Original uploaded DataFrame (unprocessed)
        predictions:    Model predictions (1D array, len = len(raw_df))
        shap_values:    2D SHAP array (n_shap_rows, n_features)
        feature_names:  List of feature names
        target_name:    Name of the target column
        shap_row_count: Number of rows SHAP was computed on (may be < len(raw_df))
    """
    buf = io.BytesIO()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:

        # ── Sheet 1: Predictions ──────────────────────────────────────────────
        pred_df = raw_df.copy()
        pred_df[f"Predicted_{target_name}"] = predictions
        pred_df.to_excel(writer, sheet_name="Predictions", index=False)

        ws1 = writer.sheets["Predictions"]
        # Header styling
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill("solid", fgColor="667EEA")
        pred_fill   = PatternFill("solid", fgColor="E8F5E9")
        bold_white  = Font(bold=True, color="FFFFFF")

        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = bold_white
            cell.alignment = Alignment(horizontal="center")

        # Highlight prediction column
        pred_col_idx = pred_df.columns.get_loc(f"Predicted_{target_name}") + 1
        for row in ws1.iter_rows(min_row=2, min_col=pred_col_idx, max_col=pred_col_idx):
            for cell in row:
                cell.fill = pred_fill

        # Auto-width
        for col in ws1.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # ── Sheet 2: Raw SHAP Values ──────────────────────────────────────────
        shap_df = pd.DataFrame(
            shap_values,
            columns=[f"SHAP_{f}" for f in feature_names]
        )
        shap_df.insert(0, "Row", range(1, len(shap_df) + 1))
        shap_df.insert(1, f"Predicted_{target_name}", predictions[:len(shap_df)])

        # Add absolute importance rank per row
        abs_shap = np.abs(shap_values)
        top_feat_per_row = [
            feature_names[int(np.argmax(abs_shap[i]))]
            for i in range(len(shap_df))
        ]
        shap_df.insert(2, "Top_Feature", top_feat_per_row)

        shap_df.to_excel(writer, sheet_name="SHAP Values", index=False)

        ws2 = writer.sheets["SHAP Values"]
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = bold_white
            cell.alignment = Alignment(horizontal="center")

        # Color positive SHAP green, negative red
        from openpyxl.styles import Font as OFont
        green_font = OFont(color="1B5E20")
        red_font   = OFont(color="B71C1C")

        shap_col_start = 4  # after Row, Prediction, Top_Feature
        for row in ws2.iter_rows(min_row=2, min_col=shap_col_start):
            for cell in row:
                try:
                    if cell.value is not None and float(cell.value) > 0:
                        cell.font = green_font
                    elif cell.value is not None and float(cell.value) < 0:
                        cell.font = red_font
                except (TypeError, ValueError):
                    pass

        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = 18

        # ── Sheet 3: Plain English Explanations ───────────────────────────────
        plain_rows = []
        for i in range(len(shap_values)):
            raw_row = raw_df.iloc[i].to_dict() if i < len(raw_df) else {}
            plain_rows.append(_plain_english_row(
                row_idx=i,
                local_shap=shap_values[i],
                feature_names=feature_names,
                raw_row=raw_row,
                prediction=predictions[i],
                top_n=3,
            ))

        plain_df = pd.DataFrame(plain_rows)
        plain_df.to_excel(writer, sheet_name="Plain English Explanations", index=False)

        ws3 = writer.sheets["Plain English Explanations"]
        reason_fill = PatternFill("solid", fgColor="FFF8E1")
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = bold_white
            cell.alignment = Alignment(horizontal="center")

        for row in ws3.iter_rows(min_row=2, min_col=3):
            for cell in row:
                cell.fill = reason_fill

        for col in ws3.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws3.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        # ── Sheet 4: Summary Stats ────────────────────────────────────────────
        mean_abs = np.abs(shap_values).mean(axis=0)
        total    = mean_abs.sum()
        summary_df = pd.DataFrame({
            "Feature":            feature_names,
            "Mean_Abs_SHAP":      mean_abs.round(5),
            "Influence_%":        [(v / total * 100) if total > 0 else 0
                                   for v in mean_abs],
            "Avg_Direction":      ["Increases" if v >= 0 else "Decreases"
                                   for v in shap_values.mean(axis=0)],
        }).sort_values("Mean_Abs_SHAP", ascending=False).reset_index(drop=True)

        summary_df["Rank"] = range(1, len(summary_df) + 1)
        summary_df = summary_df[["Rank", "Feature", "Mean_Abs_SHAP",
                                  "Influence_%", "Avg_Direction"]]
        summary_df.to_excel(writer, sheet_name="Feature Summary", index=False)

        ws4 = writer.sheets["Feature Summary"]
        for cell in ws4[1]:
            cell.fill = header_fill
            cell.font = bold_white
            cell.alignment = Alignment(horizontal="center")

        for col in ws4.columns:
            ws4.column_dimensions[col[0].column_letter].width = 22

    buf.seek(0)
    return buf.read()
